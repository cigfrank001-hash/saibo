#!/usr/bin/env python3
"""
keyword_pick.py - Stage 4: 5 词选词（强相关原则）

输入：最小意图清单 xlsx（Stage 3 输出）
半自动：对每个最小意图生成"5 平台 × 20 次"跑词 prompt
       收集 100 次结果后，自动去重 + 强相关筛选
       5 词 = 锚定词 1 + 强相关 4
输出：5 词起步包 xlsx（直接对接报价表 V1.0）
"""

import argparse
import json
import re
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUTPUT_DIR = Path(__file__).parent.parent / "output"

# 5 家 AI 平台
AI_PLATFORMS = [
    "DeepSeek",
    "豆包",
    "Kimi",
    "文心一言",
    "通义千问",
]


def build_prompts_for_intent(intent_text, brand, industry):
    """为单个最小意图生成 5 平台 × 20 次 prompt"""
    prompts = []
    for platform in AI_PLATFORMS:
        for i in range(20):
            prompts.append({
                "intent": intent_text,
                "platform": platform,
                "round": i + 1,
                "prompt": f"如果用户搜索「{intent_text}」，AI 会推荐哪些相关的搜索词？请列出 20 个不同角度的搜索词（品牌/价格/配置/对比/场景/问题等），用 JSON 数组返回。",
            })
    return prompts


def generate_all_prompts(intents, brand, industry):
    """为所有最小意图生成 prompt 池"""
    all_prompts = []
    for intent in intents:
        prompts = build_prompts_for_intent(intent, brand, industry)
        for p in prompts:
            p["intent"] = intent
            all_prompts.append(p)
    return all_prompts


def print_prompts(all_prompts):
    """打印所有 prompt"""
    print("=" * 70)
    print(f"🎯 Stage 4: 5 词选词 prompt 池")
    print("=" * 70)
    print(f"\n最小意图数：{len(set(p['intent'] for p in all_prompts))}")
    print(f"每个意图：5 平台 × 20 次 = 100 个 prompt")
    print(f"总 prompt 数：{len(all_prompts)}")
    print()

    current_intent = None
    for i, p in enumerate(all_prompts, 1):
        if p["intent"] != current_intent:
            current_intent = p["intent"]
            print("\n" + "─" * 70)
            print(f"📌 意图：{current_intent}")
            print("─" * 70)
        print(f"   [{i:03d}] {p['platform']} #{p['round']:02d}")

    print("\n" + "─" * 70)
    print(f"💡 提示：可生成完整 prompt 文件用于分发")
    print(f"   完整 prompt 模板示例：")
    sample = all_prompts[0]
    print(f"   {sample['prompt']}")


def collect_and_pick(results_file):
    """收集 100 次结果，去重 + 强相关筛选，输出 5 词起步包"""
    with open(results_file, encoding="utf-8") as f:
        results = json.load(f)

    # 按意图分组
    by_intent = {}
    for r in results:
        intent = r["intent"]
        if intent not in by_intent:
            by_intent[intent] = []
        by_intent[intent].extend(r.get("keywords", []))

    # 去重
    for intent in by_intent:
        seen = set()
        unique = []
        for kw in by_intent[intent]:
            kw_clean = kw.strip()
            if kw_clean and kw_clean.lower() not in seen:
                seen.add(kw_clean.lower())
                unique.append(kw_clean)
        by_intent[intent] = unique

    # 强相关筛选：V1.1 先去掉 [场景]/[品类]/[其他] 标签
    def clean_tag(kw):
        return re.sub(r'\s*\[[^\]]+\]$', '', kw).strip()
    
    # 合并去重：key 也清洗
    picked = {}
    for raw_intent, kws in by_intent.items():
        intent = clean_tag(raw_intent)
        # 清洗所有 kws 去标签
        kws_clean = [clean_tag(kw) for kw in kws]
        kws_clean = [k for k in kws_clean if k]  # 去空
        # 去重
        seen = set()
        unique = []
        for k in kws_clean:
            if k not in seen:
                seen.add(k)
                unique.append(k)
        # 意图本身也清洗
        intent_clean = clean_tag(intent)
        # 取包含意图核心词的优先
        intent_words = set(intent_clean.replace("推荐", "").replace("哪个好", "").split())
        scored = []
        for kw in unique:
            score = sum(1 for w in intent_words if w in kw)
            scored.append((score, kw))
        scored.sort(key=lambda x: -x[0])
        # 选 5 词：第一词 = 锚定词（=意图本身，已清洗）
        picked_kws = [intent_clean] + [kw for _, kw in scored[:4] if kw != intent_clean]
        picked[intent] = [clean_tag(kw) for kw in picked_kws[:5]]

    return picked


def write_xlsx(picked, intent_dict, out_path):
    """写 5 词起步包 xlsx（V1.1：按锚点分组多 sheet）"""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    # 样式
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    price_font = Font(name="微软雅黑", bold=True, color="C00000")
    center = Alignment(horizontal="center", vertical="center")
    brand_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    scene_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    category_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    other_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # 按锚点分组
    groups = {
        "品牌意图": {"items": [], "price": 1500, "fill": brand_fill, "label": "V0.5 拍板 ¥1,500"},
        "场景意图": {"items": [], "price": 3000, "fill": scene_fill, "label": "V1.1 新增 ¥3,000（场景锚点）"},
        "品类意图": {"items": [], "price": 3000, "fill": category_fill, "label": "V1.1 归类 ¥3,000（品类锚点）"},
        "其他意图": {"items": [], "price": 3000, "fill": other_fill, "label": "V1.1 归类 ¥3,000（其他锚点）"},
    }

    for intent, kws in picked.items():
        # V1.1：清洗 intent 标签后查 intent_dict
        intent_clean = re.sub(r'\s*\[[^\]]+\]$', '', intent).strip()
        info = intent_dict.get(intent_clean, {"type": "通意意图", "anchor": "其他"})
        if info["type"] == "品牌意图":
            groups["品牌意图"]["items"].append((intent, kws))
        elif info["anchor"] == "场景":
            groups["场景意图"]["items"].append((intent, kws))
        elif info["anchor"] == "品类":
            groups["品类意图"]["items"].append((intent, kws))
        else:
            groups["其他意图"]["items"].append((intent, kws))

    # 为每个组创建 sheet
    total_price = 0
    for sheet_name, group in groups.items():
        if not group["items"]:
            continue
        ws = wb.create_sheet(sheet_name)
        # 标题
        ws["A1"] = f"{sheet_name} - 5 词起步包（V1.1 按锚点分组）"
        ws.merge_cells("A1:G1")
        ws["A1"].font = Font(name="微软雅黑", size=14, bold=True, color="1F4E78")
        ws["A1"].alignment = center
        ws["A1"].fill = group["fill"]

        # 表头
        headers = ["#", "最小意图", "1️⃣ 锚定词", "2️⃣ 强相关", "3️⃣ 强相关", "4️⃣ 强相关", "5️⃣ 强相关"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        # 数据
        row = 3
        subtotal = 0
        for i, (intent, kws) in enumerate(group["items"], 1):
            ws.cell(row=row, column=1, value=i).alignment = center
            ws.cell(row=row, column=2, value=intent)
            for j in range(5):
                cell = ws.cell(row=row, column=3 + j, value=kws[j] if j < len(kws) else "")
                cell.alignment = center
            row += 1
            subtotal += group["price"]
            total_price += group["price"]

        row += 1
        # 小计
        ws.cell(row=row, column=1, value=f"💰 {sheet_name} 小计：{len(group['items'])} 个 × ¥{group['price']:,} = ¥{subtotal:,}").font = price_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)

        # 列宽
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 30
        for col in ["C", "D", "E", "F", "G"]:
            ws.column_dimensions[col].width = 18
        ws.freeze_panes = "A3"

    # 总计 sheet
    ws_total = wb.create_sheet("总计")
    ws_total["A1"] = "5 词起步包 - 总计（V1.1 按锚点分组）"
    ws_total.merge_cells("A1:E1")
    ws_total["A1"].font = Font(name="微软雅黑", size=14, bold=True, color="1F4E78")
    ws_total["A1"].alignment = center

    # 总计表
    for col, h in enumerate(["意图类型", "数量", "单价", "小计", "备注"], 1):
        cell = ws_total.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    row = 4
    for sheet_name, group in groups.items():
        if not group["items"]:
            continue
        ws_total.cell(row=row, column=1, value=sheet_name).font = Font(bold=True, color="1F4E78")
        ws_total.cell(row=row, column=1).fill = group["fill"]
        ws_total.cell(row=row, column=2, value=len(group["items"])).alignment = center
        ws_total.cell(row=row, column=3, value=f"¥{group['price']:,}").alignment = center
        ws_total.cell(row=row, column=4, value=f"¥{len(group['items']) * group['price']:,}").font = Font(bold=True, color="C00000")
        ws_total.cell(row=row, column=4).alignment = center
        ws_total.cell(row=row, column=5, value=group["label"])
        row += 1

    row += 1
    ws_total.cell(row=row, column=1, value=f"💰 总报价：¥{total_price:,}（直接对接报价表 V1.1）").font = Font(bold=True, color="C00000", size=12)
    ws_total.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    ws_total.cell(row=row, column=1, value=f"📌 档位：精准意图 1x（V1.1 拍板）").font = Font(italic=True, color="1F4E78")
    ws_total.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

    for col, w in zip(["A", "B", "C", "D", "E"], [16, 10, 12, 16, 30]):
        ws_total.column_dimensions[col].width = w

    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser(description="Stage 4: 5 词选词（强相关原则）")
    parser.add_argument("--input", required=True, help="Stage 3 输出（最小意图清单 xlsx）")
    parser.add_argument("--collect", action="store_true", help="收集模式")
    parser.add_argument("--results", help="收集的结果 JSON")
    args = parser.parse_args()

    # 读取 Stage 3 输出
    wb = load_workbook(args.input)
    ws = wb.active

    # 提取品牌 + 最小意图 + 锚点（V1.1：加锚点分组）
    brand = None
    intents = []  # [(intent_text, intent_type, anchor), ...]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        # 过滤空行
        if not row[1]:
            continue
        # 提取品牌（row[0] 是 "品牌"）
        if row[0] == "品牌" and row[1]:
            brand = row[1]
            continue
        # 提取最小意图（V1.1：row[9] 是 "✅ 是"，row[3] 是锚点，row[2] 是类型）
        if len(row) > 9 and row[9] == "✅ 是":
            intent_text = row[1]
            intent_type = row[2] if len(row) > 2 else "通意意图"
            anchor = row[3] if len(row) > 3 else "-"
            intents.append((intent_text, intent_type, anchor))

    if not brand:
        brand = ws["B2"].value or "未知品牌"

    print(f"📌 品牌：{brand}")
    print(f"📌 最小意图数：{len(intents)}")
    print(f"📌 最小意图：{intents[:3]}{'...' if len(intents) > 3 else ''}")

    if args.collect:
        if not args.results:
            print("❌ --collect 模式需要 --results 参数")
            return
        picked = collect_and_pick(args.results)
        # V1.1：加锚点信息到 picked（key 也去标签）
        intent_dict = {}
        for i in intents:
            intent_clean = re.sub(r'\s*\[[^\]]+\]$', '', i[0]).strip()
            intent_dict[intent_clean] = {"type": i[1], "anchor": i[2]}
        today = datetime.now().strftime("%Y%m%d")
        out_path = OUTPUT_DIR / f"5词起步包-{brand}-{today}.xlsx"
        write_xlsx(picked, intent_dict, out_path)
        print(f"\n✅ 5 词起步包已保存：{out_path}")
    else:
        # generate_all_prompts 用 (text, type, anchor) 列表
        all_prompts = generate_all_prompts([(i[0], "", "") for i in intents], brand, "")
        print_prompts(all_prompts)
        print(f"\n📝 收集 100 个结果后，运行：")
        print(f"   python3 keyword_pick.py --input {args.input} --collect --results <results.json>")


if __name__ == "__main__":
    main()
