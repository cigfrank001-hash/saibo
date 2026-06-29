#!/usr/bin/env python3
"""
intent_filter.py - Stage 3: 最小意图收口（5 维度评估）

输入：候选意图池 md（Stage 2 输出）
评估：对每个候选走 5 步判定 SOP
输出：最小意图清单 xlsx
"""

import argparse
import re
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT_DIR = Path(__file__).parent.parent / "output"

# 5 步判定 SOP 阈值
DIMENSION_THRESHOLDS = {
    "角色": {"1": "1 个身份（C 端车主）", "2+": "2+ 身份（拆/升档）"},
    "场景": {"1": "1 个情境", "2-3": "2-3 个情境", "5+": "5+ 个情境"},
    "阶段": {"决策层": "决策层", "比较层": "比较层", "认知层": "认知层"},
    "词量": {"≤10": "≤10 词（精准）", "11-30": "11-30 词（进阶）", "31-100": "31-100 词（战略）", "100+": "100+ 词（统治）"},
    "竞争": {"<3": "<3 品牌", "3-5": "3-5 品牌", "5-10": "5-10 品牌", "10+": "10+ 品牌"},
}


def parse_md_candidates(md_path):
    """解析候选意图池 md，提取 4 类意图"""
    with open(md_path, encoding="utf-8") as f:
        content = f.read()

    # 提取 frontmatter
    frontmatter = {}
    for line in content.split("\n")[:10]:
        m = re.match(r"^>\s*(\S+)\s*：\s*(.+)$", line)
        if m:
            frontmatter[m.group(1)] = m.group(2).strip()

    # 提取 4 类意图
    intent_types = ["品牌意图", "通用意图", "比较意图", "决策意图"]
    sections = {}
    for itype in intent_types:
        # 匹配 ## 标题到下一个 ## 之间的内容
        pattern = rf"## {itype}.*?\n\n(.*?)(?=\n## |\n---|\Z)"
        m = re.search(pattern, content, re.DOTALL)
        if m:
            kws = []
            for line in m.group(1).strip().split("\n"):
                # 匹配 "1. 关键词" 格式
                km = re.match(r"^\d+\.\s*(.+)$", line)
                if km:
                    kws.append(km.group(1).strip())
            sections[itype] = kws
        else:
            sections[itype] = []

    return frontmatter, sections


def evaluate_intent_5d(intent_text, intent_type, brand, competitors):
    """对单个候选意图做 5 维度评估（半自动：返回待填写模板）

    返回各维度的初始判断（基于规则），用户可调整
    """
    # 简单规则判断（实际应用中应更复杂）
    eval_result = {
        "intent_text": intent_text,
        "intent_type": intent_type,
        # 角色：1 个 = OK
        "角色": "1" if not any(kw in intent_text for kw in ["C 端", "B 端", "采购", "批发"]) else "2+",
        # 场景：数关键词
        "场景": "1",
        # 阶段：含品牌词=决策/比较层；问"什么好"=比较层；"怎么"=认知层
        "阶段": "决策层" if brand in intent_text and any(c in intent_text for c in ["价格", "配置", "续航", "参数"]) else "比较层" if "vs" in intent_text or "哪个" in intent_text or "对比" in intent_text else "认知层",
        # 词量：固定填 5（实际跑词后才知）
        "词量": "≤10",
        # 竞争：品牌意图 = <3；通意/比较 = 3+
        "竞争": "<3" if intent_type == "品牌意图" else "3-5",
    }

    # V1.1：通意意图加锚点类型
    if intent_type == "通用意图":
        if any(kw in intent_text for kw in ["夜间", "白天", "雨天", "雨雾", "长途", "接送", "通勤", "上下学", "郊游", "露营", "越野", "高速"]):
            eval_result["锚点"] = "场景"
        elif any(kw in intent_text for kw in ["30 万", "20 万", "40 万", "高端", "中端", "纯电", "混动", "SUV", "轿车", "MPV"]):
            eval_result["锚点"] = "品类"
        else:
            eval_result["锚点"] = "其他"
    else:
        eval_result["锚点"] = "-"

    # V1.1：是否最小意图单位（5 步全 OK 或 通意意图按锚点独立）
    if intent_type == "品牌意图":
        # 品牌意图：保持 5 步全 OK
        eval_result["is_minimal"] = (
            eval_result["角色"] == "1"
            and eval_result["场景"] == "1"
            and eval_result["词量"] == "≤10"
            and eval_result["竞争"] == "<3"
        )
    else:
        # 通意意图：按锚点独立成 SKU（V1.1 新增）
        # 锚点确定后，该意图的 5 词起步包是独立的可销售单位
        # 词量 ≤10 + 锚点确定 = 最小意图单位
        eval_result["is_minimal"] = (
            eval_result["词量"] == "≤10"
            and eval_result.get("锚点", "-") in ["场景", "品类", "其他"]
        )

    return eval_result


def write_xlsx(frontmatter, evaluations, out_path):
    """写最小意图清单 xlsx"""
    wb = Workbook()
    ws = wb.active
    ws.title = "最小意图清单"

    # 样式
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    # 标题
    ws["A1"] = f"最小意图清单 - {frontmatter.get('品牌', '')}"
    ws.merge_cells("A1:I1")
    ws["A1"].font = Font(name="微软雅黑", size=14, bold=True, color="1F4E78")
    ws["A1"].alignment = center

    # frontmatter
    row = 2
    for key, val in frontmatter.items():
        ws.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws.cell(row=row, column=2, value=val)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        row += 1
    row += 1

    # 表头（V1.1：加"锚点"列）
    headers = ["#", "候选意图", "类型", "锚点", "角色", "场景", "阶段", "词量", "竞争", "是否最小意图"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    row += 1

    # 数据（V1.1：加锚点列）
    minimal_count = 0
    for i, ev in enumerate(evaluations, 1):
        ws.cell(row=row, column=1, value=i).alignment = center
        ws.cell(row=row, column=2, value=ev["intent_text"])
        ws.cell(row=row, column=3, value=ev["intent_type"]).alignment = center
        anchor_cell = ws.cell(row=row, column=4, value=ev.get("锚点", "-"))
        anchor_cell.alignment = center
        # 锚点颜色：场景黄/品类蓝/其他灰
        anchor = ev.get("锚点", "-")
        if anchor == "场景":
            anchor_cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        elif anchor == "品类":
            anchor_cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        ws.cell(row=row, column=5, value=ev["角色"]).alignment = center
        ws.cell(row=row, column=6, value=ev["场景"]).alignment = center
        ws.cell(row=row, column=7, value=ev["阶段"]).alignment = center
        ws.cell(row=row, column=8, value=ev["词量"]).alignment = center
        ws.cell(row=row, column=9, value=ev["竞争"]).alignment = center
        is_minimal_cell = ws.cell(row=row, column=10, value="✅ 是" if ev["is_minimal"] else "❌ 否")
        is_minimal_cell.alignment = center
        if ev["is_minimal"]:
            is_minimal_cell.fill = ok_fill
            minimal_count += 1
        else:
            is_minimal_cell.fill = fail_fill
        row += 1

    row += 1
    # 统计（V1.1：列范围从 9 → 10）
    ws.cell(row=row, column=1, value=f"📊 统计：共 {len(evaluations)} 个候选意图，{minimal_count} 个最小意图单位").font = Font(bold=True, color="1F4E78")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1
    # V1.1：分锚点统计报价
    scene_count = sum(1 for ev in evaluations if ev.get("锚点") == "场景" and ev.get("is_minimal"))
    category_count = sum(1 for ev in evaluations if ev.get("锚点") == "品类" and ev.get("is_minimal"))
    other_count = sum(1 for ev in evaluations if ev.get("锚点") == "其他" and ev.get("is_minimal"))
    brand_count = sum(1 for ev in evaluations if ev["intent_type"] == "品牌意图" and ev.get("is_minimal"))
    total_price = brand_count * 1500 + (scene_count + category_count + other_count) * 3000
    ws.cell(row=row, column=1, value=f"💰 报价：品牌意图 {brand_count} 个 × ¥1,500 + 通意意图 {scene_count+category_count+other_count} 个 × ¥3,000 = ¥{total_price:,}")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

    # 列宽（V1.1：加锚点列）
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10  # 锚点
    for col in ["E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[col].width = 10

    # 冻结表头
    ws.freeze_panes = "A7"

    wb.save(out_path)
    print(f"✅ 最小意图清单已保存：{out_path}")
    print(f"   共 {len(evaluations)} 个候选，{minimal_count} 个最小意图单位")


def main():
    parser = argparse.ArgumentParser(description="Stage 3: 最小意图收口（5 维度评估）")
    parser.add_argument("--input", required=True, help="Stage 2 输出（候选意图池 md）")
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # 解析
    frontmatter, sections = parse_md_candidates(args.input)
    brand = frontmatter.get("品牌", "")
    competitors_str = frontmatter.get("核心竞品", "")
    competitors = [c.strip() for c in re.split(r"[ /,，、]", competitors_str) if c.strip()]

    # 评估
    evaluations = []
    for itype, kws in sections.items():
        for kw in kws:
            ev = evaluate_intent_5d(kw, itype, brand, competitors)
            evaluations.append(ev)

    # 输出
    today = datetime.now().strftime("%Y%m%d")
    out_path = OUTPUT_DIR / f"最小意图清单-{brand}-{today}.xlsx"
    write_xlsx(frontmatter, evaluations, out_path)

    print(f"\n📌 下一步：")
    print(f"   python3 keyword_pick.py --input {out_path}")


if __name__ == "__main__":
    main()
